import streamlit as st
import pandas as pd
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from difflib import SequenceMatcher

st.set_page_config(
    page_title="SuratJalan — Bulk Downloader",
    page_icon="🚛",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700;800&family=JetBrains+Mono:wght@400;500;600&display=swap');

:root {
  --bg:        #f8f7f4;
  --surface:   #ffffff;
  --surface2:  #f2f0eb;
  --border:    #e5e2d9;
  --border2:   #ccc9be;
  --text:      #1a1917;
  --text2:     #4a4641;
  --text3:     #9c9790;
  --orange:    #e55a00;
  --orange-lt: #fff4ee;
  --orange-md: #fbd0b8;
  --orange-glow: rgba(229,90,0,.12);
  --green:     #157a3c;
  --green-lt:  #edfaf3;
  --green-md:  #a8edca;
  --green-pale:#d1fae5;
  --red:       #c41c1c;
  --red-lt:    #fff0f0;
  --amber:     #b45309;
  --amber-lt:  #fffbeb;
  --amber-pale:#fef3c7;
  --blue:      #1d4ed8;
  --blue-lt:   #eff6ff;
  --blue-pale: #dbeafe;
  --violet:    #6d28d9;
  --violet-lt: #f5f3ff;
  --sky:       #0369a1;
  --sky-lt:    #f0f9ff;
  --shadow-sm: 0 1px 3px rgba(0,0,0,.06);
  --shadow:    0 4px 12px rgba(0,0,0,.08);
  --shadow-lg: 0 8px 28px rgba(0,0,0,.10);
  --radius:    14px;
  --font: 'Plus Jakarta Sans', sans-serif;
  --mono: 'JetBrains Mono', monospace;
}

*,*::before,*::after{box-sizing:border-box;}
html,body,[class*="css"]{
  font-family:var(--font)!important;
  background:var(--bg)!important;
  color:var(--text)!important;
}
.stApp{background:var(--bg)!important;}
.main .block-container{
  padding:clamp(12px,3vw,36px) clamp(12px,3vw,48px) 80px!important;
  max-width:1300px!important;
}

/* ── TOPBAR ── */
.topbar{
  display:flex;align-items:center;justify-content:space-between;
  padding:16px 0 20px;border-bottom:1.5px solid var(--border);
  margin-bottom:4px;flex-wrap:wrap;gap:10px;
}
.topbar-brand{display:flex;align-items:center;gap:12px;}
.topbar-icon{
  width:40px;height:40px;border-radius:10px;
  background:linear-gradient(135deg,#e55a00,#c44000);
  display:flex;align-items:center;justify-content:center;font-size:1.1rem;
  box-shadow:0 4px 14px rgba(229,90,0,.32);
}
.topbar-name{font-size:1.05rem;font-weight:800;color:var(--text);letter-spacing:-.3px;}
.topbar-ver{font-size:.64rem;color:var(--text3);font-weight:600;letter-spacing:.7px;text-transform:uppercase;margin-top:1px;}
.topbar-chips{display:flex;gap:6px;flex-wrap:wrap;}
.topchip{
  font-size:.68rem;font-weight:700;padding:4px 10px;
  border-radius:100px;letter-spacing:.1px;
}
.tc-orange{background:var(--orange-lt);color:var(--orange);border:1px solid var(--orange-md);}
.tc-green{background:var(--green-lt);color:var(--green);border:1px solid var(--green-md);}
.tc-blue{background:var(--blue-lt);color:var(--blue);border:1px solid #bfdbfe;}

/* ── HERO ── */
.hero{
  background:linear-gradient(135deg,#e55a00 0%,#b83c00 60%,#8b2500 100%);
  border-radius:var(--radius);
  padding:clamp(24px,4vw,36px) clamp(24px,4vw,44px);
  margin:20px 0 6px;position:relative;overflow:hidden;
  box-shadow:0 10px 32px rgba(229,90,0,.30);
}
.hero::before{
  content:'';position:absolute;top:-60px;right:-80px;
  width:280px;height:280px;
  background:rgba(255,255,255,.05);
  border-radius:50%;pointer-events:none;
}
.hero::after{
  content:'🚛';position:absolute;right:clamp(16px,4vw,36px);
  bottom:-16px;font-size:clamp(72px,14vw,120px);
  opacity:.10;pointer-events:none;transform:rotate(-6deg);
}
.hero-eyebrow{
  display:inline-flex;align-items:center;gap:7px;
  background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.2);
  border-radius:100px;padding:4px 12px;
  font-size:.65rem;font-weight:700;letter-spacing:2px;text-transform:uppercase;
  color:rgba(255,255,255,.85);margin-bottom:12px;
}
.hero-title{font-size:clamp(1.4rem,3.8vw,2.15rem);font-weight:800;color:#fff;line-height:1.12;margin-bottom:10px;letter-spacing:-.5px;}
.hero-title span{opacity:.7;}
.hero-desc{font-size:clamp(.77rem,1.8vw,.9rem);color:rgba(255,255,255,.72);line-height:1.75;max-width:560px;}
.hero-stats{display:flex;gap:20px;margin-top:18px;flex-wrap:wrap;}
.hero-stat{border-left:2px solid rgba(255,255,255,.25);padding-left:12px;}
.hero-stat-val{font-family:var(--mono);font-size:1.1rem;font-weight:700;color:#fff;}
.hero-stat-lbl{font-size:.62rem;color:rgba(255,255,255,.55);text-transform:uppercase;letter-spacing:1px;margin-top:1px;}

/* ── STEP PILLS ── */
.steps{
  display:flex;align-items:center;gap:0;
  background:var(--surface);border:1.5px solid var(--border);
  border-radius:100px;padding:5px 8px;margin:20px 0;
  box-shadow:var(--shadow-sm);overflow-x:auto;
  width:fit-content;max-width:100%;
}
.step-pill{
  display:flex;align-items:center;gap:7px;
  padding:6px 14px;border-radius:100px;
  font-size:.76rem;font-weight:700;white-space:nowrap;
  transition:all .2s;
}
.step-pill.done{background:var(--green-lt);color:var(--green);}
.step-pill.active{background:var(--orange);color:#fff;box-shadow:0 2px 10px rgba(229,90,0,.30);}
.step-pill.idle{color:var(--text3);}
.step-num{
  width:20px;height:20px;border-radius:50%;
  font-size:.65rem;font-weight:800;
  display:flex;align-items:center;justify-content:center;flex-shrink:0;
}
.step-pill.done .step-num{background:var(--green);color:#fff;}
.step-pill.active .step-num{background:rgba(255,255,255,.25);color:#fff;}
.step-pill.idle .step-num{background:var(--surface2);color:var(--text3);}
.step-arrow{color:var(--border2);font-size:.7rem;padding:0 3px;flex-shrink:0;}

/* ── SECTION HEADER ── */
.sec-hdr{
  display:flex;align-items:center;gap:10px;
  margin:clamp(24px,4vw,36px) 0 clamp(14px,2vw,20px);
}
.sec-badge{
  background:var(--orange);color:#fff;
  font-size:.64rem;font-weight:800;letter-spacing:1.5px;
  text-transform:uppercase;padding:4px 10px;border-radius:6px;
  box-shadow:0 2px 8px rgba(229,90,0,.24);
}
.sec-title{font-size:.98rem;font-weight:700;color:var(--text);letter-spacing:-.2px;}
.sec-line{flex:1;height:1.5px;background:linear-gradient(to right, var(--border), transparent);}

/* ── UPLOAD CARDS ── */
.ucard{
  background:var(--surface);border:1.5px solid var(--border);
  border-radius:var(--radius);padding:clamp(16px,2.5vw,22px);
  box-shadow:var(--shadow-sm);transition:all .22s;
  position:relative;overflow:hidden;
}
.ucard::before{
  content:'';position:absolute;top:0;left:0;right:0;height:3px;
  border-radius:2px 2px 0 0;background:var(--border);
  transition:background .22s;
}
.ucard:hover{border-color:var(--orange-md);box-shadow:var(--shadow);}
.ucard:hover::before{background:linear-gradient(90deg,var(--orange),#f97316);}
.ucard.ready{border-color:#6ee7b7;background:linear-gradient(160deg,var(--surface) 60%,var(--green-lt));}
.ucard.ready::before{background:linear-gradient(90deg,var(--green),#22c55e);}
.ucard-head{display:flex;align-items:flex-start;gap:12px;margin-bottom:12px;}
.ucard-icon{
  width:44px;height:44px;border-radius:11px;flex-shrink:0;
  display:flex;align-items:center;justify-content:center;font-size:1.15rem;
}
.ucard-icon-a{background:var(--orange-lt);border:1.5px solid var(--orange-md);}
.ucard-icon-b{background:var(--green-lt);border:1.5px solid var(--green-md);}
.ucard-ttl{font-size:.9rem;font-weight:700;color:var(--text);}
.ucard-sub{font-size:.72rem;color:var(--text3);margin-top:3px;line-height:1.5;}
.tag-row{display:flex;flex-wrap:wrap;gap:5px;margin-top:8px;}
.tag-r{font-size:.63rem;font-weight:700;padding:3px 8px;border-radius:5px;font-family:var(--mono);
  background:var(--orange-lt);color:var(--orange);border:1px solid var(--orange-md);}
.tag-o{font-size:.63rem;font-weight:700;padding:3px 8px;border-radius:5px;font-family:var(--mono);
  background:var(--surface2);color:var(--text3);border:1px solid var(--border);}
.file-pill{
  display:flex;align-items:center;gap:8px;
  background:var(--green-lt);border:1px solid #6ee7b7;
  border-radius:8px;padding:7px 12px;margin-top:10px;
  font-size:.77rem;color:var(--green);font-weight:600;
}

/* ── BUTTONS ── */
.stButton>button{
  font-family:var(--font)!important;font-weight:600!important;
  font-size:.82rem!important;border-radius:9px!important;
  padding:9px 18px!important;transition:all .18s!important;letter-spacing:-.1px!important;
}
.stButton>button[kind="primary"]{
  background:linear-gradient(135deg,#e55a00,#c44000)!important;color:#fff!important;
  border:none!important;
  box-shadow:0 3px 14px rgba(229,90,0,.32)!important;
}
.stButton>button[kind="primary"]:hover{
  transform:translateY(-1px)!important;
  box-shadow:0 6px 20px rgba(229,90,0,.42)!important;
}
.stButton>button[kind="primary"]:disabled{
  background:var(--border2)!important;color:var(--text3)!important;
  box-shadow:none!important;transform:none!important;
}
.stButton>button:not([kind="primary"]){
  background:var(--surface)!important;color:var(--text2)!important;
  border:1.5px solid var(--border)!important;box-shadow:var(--shadow-sm)!important;
}
.stButton>button:not([kind="primary"]):hover{
  border-color:var(--orange-md)!important;color:var(--orange)!important;
  background:var(--orange-lt)!important;transform:translateY(-1px)!important;
}

/* ── STAT CARDS ── */
.stat-grid{
  display:grid;
  grid-template-columns:repeat(auto-fit,minmax(min(155px,100%),1fr));
  gap:clamp(8px,1.5vw,12px);margin:clamp(14px,2.5vw,22px) 0;
}
.stat-card{
  background:var(--surface);border:1.5px solid var(--border);
  border-radius:var(--radius);padding:clamp(14px,2vw,20px);
  box-shadow:var(--shadow-sm);position:relative;overflow:hidden;
  transition:box-shadow .18s,transform .18s;cursor:default;
}
.stat-card:hover{box-shadow:var(--shadow);transform:translateY(-2px);}
.stat-stripe{position:absolute;top:0;left:0;right:0;height:3px;border-radius:2px 2px 0 0;}
.stat-icon-bg{position:absolute;right:-4px;bottom:-8px;font-size:2.8rem;opacity:.055;pointer-events:none;}
.stat-lbl{font-size:.62rem;font-weight:700;text-transform:uppercase;letter-spacing:1.2px;color:var(--text3);margin-bottom:6px;}
.stat-val{font-family:var(--mono);font-size:clamp(1.6rem,3.5vw,2.3rem);font-weight:600;line-height:1;}
.stat-sub{font-size:.66rem;font-weight:600;margin-top:5px;}

/* ── ALERTS ── */
.alert{
  border-radius:10px;padding:clamp(10px,2vw,13px) clamp(12px,2vw,16px);
  margin:8px 0;font-size:.82rem;line-height:1.65;
  display:flex;gap:11px;align-items:flex-start;
}
.alert-ico{font-size:.9rem;flex-shrink:0;margin-top:2px;}
.a-success{background:var(--green-lt);border:1px solid #a8edca;border-left:3px solid var(--green);color:#0e5228;}
.a-warn{background:var(--amber-lt);border:1px solid #fde68a;border-left:3px solid var(--amber);color:#78350f;}
.a-error{background:var(--red-lt);border:1px solid #fca5a5;border-left:3px solid var(--red);color:#7f1d1d;}
.a-info{background:var(--blue-lt);border:1px solid #bfdbfe;border-left:3px solid var(--blue);color:#1e3a8a;}
.a-violet{background:var(--violet-lt);border:1px solid #ddd6fe;border-left:3px solid var(--violet);color:#3b0764;}
.a-sky{background:var(--sky-lt);border:1px solid #bae6fd;border-left:3px solid var(--sky);color:#0c4a6e;}
.a-orange{background:var(--orange-lt);border:1px solid var(--orange-md);border-left:3px solid var(--orange);color:#7c2d12;}

/* ── TABS ── */
.stTabs [data-baseweb="tab-list"]{
  background:var(--surface2)!important;border:1.5px solid var(--border)!important;
  border-radius:12px!important;padding:4px!important;gap:2px!important;
  box-shadow:var(--shadow-sm)!important;overflow-x:auto!important;flex-wrap:nowrap!important;
}
.stTabs [data-baseweb="tab"]{
  border-radius:8px!important;font-size:clamp(.69rem,1.6vw,.79rem)!important;
  font-weight:600!important;padding:clamp(6px,1.5vw,9px) clamp(9px,1.8vw,14px)!important;
  color:var(--text3)!important;border:none!important;background:transparent!important;
  white-space:nowrap!important;font-family:var(--font)!important;transition:all .18s!important;
}
.stTabs [data-baseweb="tab"]:hover{color:var(--text2)!important;background:rgba(255,255,255,.6)!important;}
.stTabs [aria-selected="true"]{
  background:var(--surface)!important;color:var(--orange)!important;
  font-weight:700!important;box-shadow:var(--shadow-sm)!important;
}
.stTabs [data-baseweb="tab-highlight"],.stTabs [data-baseweb="tab-border"]{display:none!important;}

/* ── PROGRESS ── */
.stProgress>div>div>div>div{background:linear-gradient(90deg,var(--orange),#f97316)!important;border-radius:100px!important;}
.stProgress>div>div{background:var(--surface2)!important;border-radius:100px!important;}

/* ── INPUT ── */
.stTextInput>div>div>input{
  background:var(--surface)!important;border:1.5px solid var(--border)!important;
  border-radius:9px!important;color:var(--text)!important;
  font-family:var(--font)!important;font-size:.83rem!important;
  padding:9px 13px!important;transition:border-color .18s,box-shadow .18s!important;
  box-shadow:var(--shadow-sm)!important;
}
.stTextInput>div>div>input:focus{border-color:var(--orange)!important;box-shadow:0 0 0 3px rgba(229,90,0,.10)!important;}
.stTextInput>div>div>input::placeholder{color:var(--text3)!important;}

/* ── SLIDER ── */
.stSlider>div>div>div>div{background:var(--orange)!important;}
.stSlider>div>div>div{background:var(--border)!important;}

/* ── DATAFRAME ── */
div[data-testid="stDataFrame"]>div{
  background:var(--surface)!important;border:1.5px solid var(--border)!important;
  border-radius:12px!important;box-shadow:var(--shadow-sm)!important;
}

/* ── DOWNLOAD BUTTONS ── */
.stDownloadButton>button{
  font-family:var(--font)!important;font-weight:700!important;font-size:.79rem!important;
  background:var(--green-lt)!important;border:1.5px solid #6ee7b7!important;
  border-radius:9px!important;color:var(--green)!important;
  padding:8px 14px!important;transition:all .18s!important;box-shadow:var(--shadow-sm)!important;
}
.stDownloadButton>button:hover{
  background:var(--green-md)!important;border-color:#34d399!important;
  box-shadow:0 4px 12px rgba(21,122,60,.22)!important;transform:translateY(-1px)!important;
}

/* ── FILE UPLOADER ── */
div[data-testid="stFileUploaderDropzone"]{
  background:var(--surface2)!important;border:1.5px dashed var(--border2)!important;
  border-radius:10px!important;transition:all .2s!important;
}
div[data-testid="stFileUploaderDropzone"]:hover{
  border-color:var(--orange-md)!important;background:var(--orange-lt)!important;
}
div[data-testid="stFileUploaderDropzone"] p{
  color:var(--text3)!important;font-size:.79rem!important;font-family:var(--font)!important;
}
button[data-testid="baseButton-secondary"]{
  background:var(--orange-lt)!important;border:1.5px solid var(--orange-md)!important;
  color:var(--orange)!important;border-radius:8px!important;font-family:var(--font)!important;
}

/* ── CHECKBOX ── */
.stCheckbox > label {
  font-family: var(--font) !important;
  font-size: .82rem !important;
  color: var(--text2) !important;
  cursor: pointer !important;
}
.stCheckbox > label > span[data-testid="stCheckboxLabel"] {
  font-size: .82rem !important;
  font-family: var(--font) !important;
}

/* ── NOPOL PILLS ── */
.np{
  display:inline-block;background:var(--orange-lt);border:1px solid var(--orange-md);
  color:var(--orange);border-radius:6px;padding:2px 9px;
  font-family:var(--mono);font-size:.79rem;font-weight:500;
}
.np-dup{background:var(--violet-lt);border-color:#c4b5fd;color:var(--violet);}

/* ── SIM BADGES ── */
.sim{display:inline-block;border-radius:6px;padding:3px 8px;font-size:.71rem;font-weight:700;font-family:var(--mono);}
.sim-hi{background:var(--green-lt);color:var(--green);border:1px solid var(--green-md);}
.sim-md{background:var(--amber-lt);color:var(--amber);border:1px solid #fde68a;}
.sim-lo{background:var(--red-lt);color:var(--red);border:1px solid #fca5a5;}

/* ── SARAN ROW ── */
.saran-row{
  background:var(--surface);border:1.5px solid var(--border);
  border-radius:11px;padding:12px 16px;margin:6px 0;
  transition:border-color .18s,box-shadow .18s;
  position:relative;
}
.saran-row:hover{border-color:var(--orange-md);box-shadow:var(--shadow-sm);}
.saran-row.selected{border-color:#6ee7b7;background:var(--green-lt);}
.saran-row.selected::before{
  content:'✓';position:absolute;top:8px;right:10px;
  background:var(--green);color:#fff;font-size:.64rem;font-weight:800;
  width:18px;height:18px;border-radius:50%;display:flex;align-items:center;justify-content:center;
  line-height:18px;text-align:center;
}

/* ── DUP GROUP ── */
.dup-group{
  background:var(--violet-lt);border:1.5px solid #ddd6fe;
  border-radius:var(--radius);padding:14px 18px;margin:12px 0 6px;
}
.dup-group-header{display:flex;align-items:center;gap:9px;flex-wrap:wrap;margin-bottom:12px;}

/* ── BULK ACTION BAR ── */
.bulk-bar{
  background:linear-gradient(135deg,var(--green-lt),#f0fdf4);
  border:1.5px solid #6ee7b7;border-radius:11px;
  padding:12px 18px;margin:12px 0;
  display:flex;align-items:center;justify-content:space-between;
  flex-wrap:wrap;gap:10px;
  box-shadow:0 2px 8px rgba(21,122,60,.10);
}
.bulk-bar-info{font-size:.83rem;font-weight:700;color:var(--green);}
.bulk-bar-info span{font-family:var(--mono);}
.bulk-bar-sub{font-size:.72rem;color:#4ade80;margin-top:2px;}

/* ── SELECT CONTROLS ── */
.sel-controls{
  display:flex;align-items:center;gap:8px;
  background:var(--surface);border:1.5px solid var(--border);
  border-radius:9px;padding:8px 14px;margin:10px 0;
  box-shadow:var(--shadow-sm);
}
.sel-label{font-size:.76rem;font-weight:700;color:var(--text2);}
.sel-count{
  background:var(--orange-lt);color:var(--orange);border:1px solid var(--orange-md);
  border-radius:100px;padding:2px 9px;font-size:.7rem;font-weight:700;
  font-family:var(--mono);
}

/* ── TABLE ── */
.col-hdr{font-size:.62rem;font-weight:700;text-transform:uppercase;letter-spacing:1.2px;color:var(--text3);}
.row-div{border:none;border-top:1px solid var(--border);margin:8px 0;}
.cache-info{
  background:var(--surface2);border:1px solid var(--border);
  border-radius:9px;padding:8px 14px;font-size:.77rem;color:var(--text2);line-height:1.4;
}

/* ── EMPTY STATE ── */
.empty-state{
  text-align:center;padding:clamp(36px,5vw,60px) 24px;
  color:var(--text3);background:var(--surface);
  border:1.5px dashed var(--border);border-radius:var(--radius);margin:10px 0;
}
.empty-state .ico{font-size:clamp(2rem,4vw,2.8rem);margin-bottom:12px;opacity:.4;}
.empty-state h3{font-size:.98rem;font-weight:700;color:var(--text2);margin-bottom:6px;}
.empty-state p{font-size:.82rem;line-height:1.65;max-width:320px;margin:0 auto;}

/* ── SIDEBAR ── */
section[data-testid="stSidebar"]{background:var(--surface)!important;border-right:1.5px solid var(--border)!important;}
section[data-testid="stSidebar"] *{font-family:var(--font)!important;color:var(--text2)!important;}
section[data-testid="stSidebar"] h2,section[data-testid="stSidebar"] h3{color:var(--text)!important;}

/* ── EXPANDER ── */
details{background:var(--surface)!important;border-radius:10px!important;border:1.5px solid var(--border)!important;}
summary{font-size:.81rem!important;color:var(--text2)!important;padding:10px 14px!important;font-family:var(--font)!important;}

/* ── SCROLLBAR ── */
::-webkit-scrollbar{width:4px;height:4px;}
::-webkit-scrollbar-track{background:var(--surface2);}
::-webkit-scrollbar-thumb{background:var(--border2);border-radius:100px;}
::-webkit-scrollbar-thumb:hover{background:var(--text3);}

#MainMenu,footer,header{visibility:hidden;}

@media(max-width:640px){
  .stat-grid{grid-template-columns:repeat(2,1fr);}
  .topbar-chips{display:none;}
  .hero::after{display:none;}
  .hero-stats{display:none;}
}
@media(max-width:400px){
  .stat-grid{grid-template-columns:1fr 1fr;}
}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# UTILITIES
# ══════════════════════════════════════════════════════════════════════════════
def norm_nopol(v):
    if v is None or (isinstance(v, float) and pd.isna(v)): return ''
    s = str(v).strip().upper()
    s = re.sub(r'\s+', ' ', s)
    s = re.sub(r'([A-Z])(\d)', r'\1 \2', s)
    s = re.sub(r'(\d)([A-Z])', r'\1 \2', s)
    return re.sub(r'\s+', ' ', s).strip()

def norm_kuantum(v):
    try: return round(float(str(v).replace(',', '.').strip()), 3)
    except: return None

def find_col(df, kws):
    for col in df.columns:
        c = col.lower().replace(' ', '').replace('_', '')
        for k in kws:
            if k.replace(' ', '').replace('_', '') in c: return col
    return None

def extract_fid(link):
    if not isinstance(link, str) or not link.strip(): return None
    for p in [r'/file/d/([a-zA-Z0-9_-]+)', r'id=([a-zA-Z0-9_-]+)',
              r'/d/([a-zA-Z0-9_-]+)', r'open\?id=([a-zA-Z0-9_-]+)']:
        m = re.search(p, link.strip())
        if m and len(m.group(1)) >= 15: return m.group(1)
    return None

def to_preview(link):
    fid = extract_fid(link)
    return f'https://drive.google.com/file/d/{fid}/preview' if fid else None

def detect_file_type(content):
    if not content or len(content) < 4: return 'unknown'
    sig = content[:8]
    if sig[:4] == b'%PDF': return 'pdf'
    if sig[:3] == b'\xff\xd8\xff': return 'jpg'
    if sig[:8] == b'\x89PNG\r\n\x1a\n': return 'png'
    try:
        snippet = content[:2000].decode('utf-8', errors='ignore').lower()
        if '<html' in snippet or '<!doctype' in snippet: return 'html'
    except: pass
    return 'unknown'

def _get_bytes(resp): return b''.join(resp.iter_content(chunk_size=65536))

def download_gdrive(fid, retries=4, timeout=45):
    if not fid: return None
    session = requests.Session()
    session.headers.update({'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'})
    base_url = f'https://drive.google.com/uc?export=download&id={fid}'
    for attempt in range(retries):
        try:
            r1 = session.get(base_url, timeout=timeout, stream=True, allow_redirects=True)
            raw1 = _get_bytes(r1)
            if 'text/html' not in r1.headers.get('Content-Type', ''):
                if detect_file_type(raw1) != 'html' and len(raw1) > 512: return raw1
            html = raw1.decode('utf-8', errors='ignore')
            m = re.search(r'name="confirm"\s+value="([^"]+)"', html)
            if m:
                r2 = session.get(f'{base_url}&confirm={m.group(1)}', timeout=timeout, stream=True, allow_redirects=True)
                raw2 = _get_bytes(r2)
                if detect_file_type(raw2) != 'html' and len(raw2) > 512: return raw2
            for extra in ['&confirm=t', '']:
                r = session.get(f'{base_url}{extra}', timeout=timeout, stream=True, allow_redirects=True)
                raw = _get_bytes(r)
                if detect_file_type(raw) != 'html' and len(raw) > 512: return raw
        except: pass
        if attempt < retries - 1: time.sleep(2 ** attempt)
    return None

def download_file(link):
    if not isinstance(link, str) or not link.strip(): return None
    fid = extract_fid(link)
    if fid: return download_gdrive(fid)
    try:
        r = requests.get(link.strip(), timeout=45, stream=True, headers={'User-Agent': 'Mozilla/5.0'})
        raw = _get_bytes(r)
        if detect_file_type(raw) != 'html' and len(raw) > 512: return raw
    except: pass
    return None

def infer_extension(content, fallback='pdf'):
    return {'pdf':'pdf','jpg':'jpg','png':'png'}.get(detect_file_type(content) if content else 'x', fallback)

def make_safe_filename(nopol, kuantum, idx, ext, total=999, dup_label=''):
    safe = re.sub(r'[\\/:*?"<>|]', '_', str(nopol))
    pad = len(str(max(total, 1)))
    no = str(idx + 1).zfill(pad)
    return f'{no}_{safe}_{kuantum}{dup_label}.{ext}'

def make_zip(files):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for name, data in files.items(): zf.writestr(name, data)
    buf.seek(0); return buf.read()

def img_bytes_to_pdf(img_bytes):
    tmp_path = None
    try:
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.pagesizes import A4
        from PIL import Image as PILImage
        pil = PILImage.open(io.BytesIO(img_bytes))
        if pil.mode not in ('RGB', 'L'): pil = pil.convert('RGB')
        w_px, h_px = pil.size; a4_w, a4_h = A4; margin = 20
        scale = min((a4_w - 2 * margin) / w_px, (a4_h - 2 * margin) / h_px)
        draw_w, draw_h = w_px * scale, h_px * scale
        x = (a4_w - draw_w) / 2; y = (a4_h - draw_h) / 2
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
            tmp_path = tmp.name; pil.save(tmp_path, format='PNG')
        pdf_buf = io.BytesIO()
        c = rl_canvas.Canvas(pdf_buf, pagesize=(a4_w, a4_h))
        c.drawImage(tmp_path, x, y, width=draw_w, height=draw_h, preserveAspectRatio=True)
        c.save(); pdf_buf.seek(0); return pdf_buf.read()
    except: return None
    finally:
        if tmp_path:
            try: os.unlink(tmp_path)
            except: pass

def merge_pdfs(content_list):
    try: from pypdf import PdfWriter, PdfReader
    except ImportError:
        try: from PyPDF2 import PdfWriter, PdfReader
        except ImportError: return None
    writer = PdfWriter()
    for ct in content_list:
        if not ct: continue
        ftype = detect_file_type(ct)
        pdf_data = ct if ftype == 'pdf' else (img_bytes_to_pdf(ct) if ftype in ('jpg', 'png') else None)
        if pdf_data:
            try:
                for page in PdfReader(io.BytesIO(pdf_data)).pages: writer.add_page(page)
            except: continue
    if len(writer.pages) == 0: return None
    out = io.BytesIO(); writer.write(out); out.seek(0); return out.read()

# ══════════════════════════════════════════════════════════════════════════════
# FILE READING — HYPERLINK-AWARE
# ══════════════════════════════════════════════════════════════════════════════
def read_file(f):
    if f.name.lower().endswith('.csv'):
        return pd.read_csv(f)
    return pd.read_excel(f)

def read_xlsx_with_hyperlinks(f):
    try:
        from openpyxl import load_workbook
        f.seek(0)
        wb = load_workbook(f, data_only=True)
        ws = wb.active
        hyperlinks = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.hyperlink:
                    target = cell.hyperlink if isinstance(cell.hyperlink, str) else cell.hyperlink.target
                    if target:
                        hyperlinks[(cell.row, cell.column)] = target
        if not hyperlinks:
            f.seek(0); return pd.read_excel(f)
        f.seek(0); df = pd.read_excel(f)
        for (row, col), target in hyperlinks.items():
            if row < 2: continue
            pandas_row = row - 2; pandas_col = col - 1
            if pandas_row < len(df) and pandas_col < len(df.columns):
                current_val = str(df.iloc[pandas_row, pandas_col]).strip()
                if not current_val.startswith('http'):
                    df.iloc[pandas_row, pandas_col] = target
        return df
    except Exception:
        f.seek(0); return pd.read_excel(f)

def load_file1(df):
    nc = find_col(df, ['nopol','nomor polisi','no pol','no.pol','nopolisi']) or find_col(df, ['pol'])
    kc = find_col(df, ['kuantum','quantum','tonase','tonage','qty','jumlah','volume','berat'])
    if not nc:
        st.error(f"❌ Kolom NOPOL tidak ditemukan. Kolom: `{'`, `'.join(df.columns)}`")
        return pd.DataFrame()
    if not kc:
        st.error("❌ Kolom KUANTUM tidak ditemukan."); return pd.DataFrame()
    out = pd.DataFrame()
    out['nopol'] = df[nc].apply(norm_nopol)
    out['kuantum'] = df[kc].apply(norm_kuantum)
    out = out[(out['nopol'] != '') & out['nopol'].notna()].dropna(subset=['kuantum'])
    return out[out['kuantum'] > 0].reset_index(drop=True)

def load_file2(raw_df):
    df = raw_df.copy()
    if df.empty: return pd.DataFrame()
    frow = df.iloc[0].tolist()
    if any(str(v).upper().strip() in ['NOPOL','KUANTUM','FOTO SURAT JALAN','SURAT JALAN'] for v in frow):
        df.columns = [str(c).strip() for c in df.iloc[0]]; df = df[1:].reset_index(drop=True)
    nc = find_col(df, ['nopol','nomor polisi','no pol','no truk','no.pol','nopolisi']) or find_col(df, ['pol'])
    kc = find_col(df, ['kuantum','quantum','tonase','tonage','qty','jumlah','volume','berat'])
    lc = find_col(df, ['surat jalan','suratjalan','foto surat','foto','link','url','drive','gdrive'])
    if not nc: st.error("❌ Kolom NOPOL tidak ditemukan di File 2."); return pd.DataFrame()
    if not kc: st.error("❌ Kolom KUANTUM tidak ditemukan di File 2."); return pd.DataFrame()
    if not lc: st.error("❌ Kolom SURAT JALAN / LINK tidak ditemukan di File 2."); return pd.DataFrame()
    out = pd.DataFrame()
    out['nopol'] = df[nc].apply(lambda x: norm_nopol(str(x)) if pd.notna(x) else '')
    out['kuantum'] = df[kc].apply(norm_kuantum)
    out['surat_jalan'] = df[lc].astype(str).str.strip()
    out = out[out['nopol'] != '']
    out = out[out['nopol'].str.upper() != 'NOPOL']
    out = out.dropna(subset=['kuantum']); out = out[out['kuantum'] > 0]
    valid = (
        out['surat_jalan'].str.startswith('http') |
        out['surat_jalan'].str.lower().str.endswith('.jpg') |
        out['surat_jalan'].str.lower().str.endswith('.pdf') |
        out['surat_jalan'].str.lower().str.endswith('.png')
    )
    return out[valid].reset_index(drop=True)

def nopol_similarity(a, b): return SequenceMatcher(None, a, b).ratio()

def find_nopol_suggestions(nopol_f1, kuantum, df2, top_n=5, min_similarity=0.5):
    same_kuantum = df2[df2['kuantum'] == kuantum].copy()
    if same_kuantum.empty: return []
    results = []
    for _, row in same_kuantum.iterrows():
        sim = nopol_similarity(nopol_f1, row['nopol'])
        if sim >= min_similarity:
            results.append({'nopol_f2': row['nopol'], 'kuantum': int(row['kuantum']),
                            'surat_jalan': row['surat_jalan'], 'similarity': round(sim * 100, 1)})
    results.sort(key=lambda x: x['similarity'], reverse=True)
    return results[:top_n]

def build_missing_with_suggestions(missing_df, df1, df2):
    rows = []
    for _, row in missing_df.iterrows():
        nopol = row['nopol']; kuantum = int(row['kuantum'])
        f2_match = df2[df2['nopol'] == nopol]
        if len(f2_match) > 0:
            ks = sorted(f2_match['kuantum'].dropna().astype(int).unique().tolist())
            d = ', '.join(map(str, ks[:8])) + (f' (+{len(ks)-8})' if len(ks) > 8 else '')
            rows.append({'nopol': nopol, 'kuantum': kuantum, 'kategori': 'kuantum_beda', 'info': d, 'saran': []})
        else:
            saran = find_nopol_suggestions(nopol, kuantum, df2)
            rows.append({'nopol': nopol, 'kuantum': kuantum,
                         'kategori': 'nopol_mirip' if saran else 'tidak_ada', 'info': '', 'saran': saran})
    return rows

def detect_duplicates_f1(df1):
    key = ['nopol', 'kuantum']
    counts = df1.groupby(key).size().reset_index(name='jumlah_duplikat')
    dup_keys = counts[counts['jumlah_duplikat'] > 1][key]
    if dup_keys.empty: return pd.DataFrame()
    merged = df1.merge(dup_keys, on=key, how='inner')
    merged = merged.merge(counts, on=key, how='left')
    merged['baris_ke'] = merged.groupby(key).cumcount() + 1
    return merged.reset_index(drop=True)

def match_files(df1, df2):
    df2_valid = df2[df2['surat_jalan'].str.startswith('http', na=False)].copy()
    pos_in_group = df1.groupby(['nopol', 'kuantum']).cumcount()
    result_rows = []
    for idx, row1 in df1.iterrows():
        pos = int(pos_in_group[idx])
        matches = df2_valid[(df2_valid['nopol'] == row1['nopol']) & (df2_valid['kuantum'] == row1['kuantum'])].reset_index(drop=True)
        if len(matches) > 0:
            mrow = matches.iloc[min(pos, len(matches) - 1)]
            result_rows.append({'nopol': row1['nopol'], 'kuantum': row1['kuantum'],
                                 'surat_jalan': mrow['surat_jalan'], '_f1_idx': idx, '_link_no': pos + 1})
        else:
            result_rows.append({'nopol': row1['nopol'], 'kuantum': row1['kuantum'],
                                 'surat_jalan': None, '_f1_idx': idx, '_link_no': 0})
    return pd.DataFrame(result_rows)

def _worker(task, cache_snapshot):
    link = task['link']
    ct = cache_snapshot.get(link)
    if ct is None: ct = download_file(link)
    return {**task, 'content': ct}

def run_bulk_download(rows, label=''):
    cache_snapshot = dict(st.session_state.dl_cache)
    tasks = [{'idx': r['idx'], 'nopol': r['nopol'], 'kuantum': r['kuantum'],
               'link': r['link'], 'dup_label': r.get('dup_label', '')} for r in rows]
    prog_container = st.empty()
    with prog_container.container():
        prog_bar = st.progress(0)
        st.markdown(f'<div class="alert a-info"><span class="alert-ico">⏳</span><div><b>Mengunduh {label}…</b> Proses paralel 8 thread, harap tunggu.</div></div>', unsafe_allow_html=True)
        status_txt = st.empty()
    ok_files = {}; new_cache = {}; fail_list = []; done_n = 0; total = len(tasks)
    with ThreadPoolExecutor(max_workers=8) as ex:
        futs = {ex.submit(_worker, t, cache_snapshot): t for t in tasks}
        for fut in as_completed(futs):
            res = fut.result(); ct = res['content']
            if ct:
                new_cache[res['link']] = ct
                ext = infer_extension(ct)
                fn = make_safe_filename(res['nopol'], res['kuantum'], res['idx'], ext, total=total, dup_label=res['dup_label'])
                base, c = fn, 1
                while fn in ok_files:
                    fn = base.rsplit('.', 1)[0] + f'_{c}.' + base.rsplit('.', 1)[-1]; c += 1
                ok_files[fn] = ct
            else:
                fail_list.append(f"{res['nopol']} ({res['kuantum']}){res['dup_label']}")
            done_n += 1; prog_bar.progress(done_n / total)
            pct = int(done_n / total * 100)
            status_txt.markdown(f"**{pct}%** — {done_n}/{total} &nbsp;|&nbsp; ✅ **{len(ok_files)}** berhasil &nbsp;|&nbsp; ❌ **{len(fail_list)}** gagal")
    prog_container.empty()
    return ok_files, fail_list, new_cache

# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════
for _k in ['result_df','missing_df','nopol_diff_df','nopol_miss_df','active_preview',
           'df2_debug','df1_debug','dl_cache','dup_df','missing_detail',
           'saran_preview','dup_prev_active','processed','saran_selected']:
    if _k not in st.session_state: st.session_state[_k] = None
if st.session_state.dl_cache is None: st.session_state.dl_cache = {}
if st.session_state.saran_preview is None: st.session_state.saran_preview = {}
if st.session_state.dup_prev_active is None: st.session_state.dup_prev_active = {}
if st.session_state.processed is None: st.session_state.processed = False
# saran_selected: dict of {saran_key: bool} for checkboxes in Tab 3
if st.session_state.saran_selected is None: st.session_state.saran_selected = {}

# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown('''<div style="text-align:center;padding:20px 0 14px">
      <div style="width:50px;height:50px;background:linear-gradient(135deg,#e55a00,#c44000);border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:1.4rem;margin:0 auto 12px;box-shadow:0 4px 14px rgba(229,90,0,.30)">🚛</div>
      <div style="font-weight:800;font-size:1rem;color:#1a1917">SuratJalan</div>
      <div style="font-size:.65rem;color:#9c9790;margin-top:3px;text-transform:uppercase;letter-spacing:1px">Bulk Downloader v2.2</div>
    </div>''', unsafe_allow_html=True)
    st.divider()
    with st.expander("📋 Cara Penggunaan", expanded=True):
        st.markdown("1. **Upload File 1** — Daftar target (NOPOL + Kuantum)\n2. **Upload File 2** — Database link Google Drive\n3. Klik **⚙️ Proses & Cocokkan**\n4. Download **ZIP** atau **PDF Gabungan**\n5. **Tab NOPOL Tidak Ada** — Pilih saran & download bulk")
    with st.expander("📂 Format File"):
        st.markdown("**File 1:** `NOPOL` + `KUANTUM`\n\n**File 2:** `NOPOL` + `KUANTUM` + `Link GDrive`\n\n✅ Mendukung **hyperlink** (cell berisi teks seperti 'Link' dengan URL tersembunyi)\n\nFormat: `.xlsx`, `.xls`, `.csv`")
    with st.expander("❓ FAQ"):
        st.markdown("**Link di Excel berupa hyperlink?** Otomatis terdeteksi!\n\n**File gagal?** Mungkin private/expired.\n\n**Duplikat?** NOPOL+Kuantum muncul >1x.\n\n**Tab NOPOL Tidak Ada?** Centang saran yang ingin didownload, lalu klik ZIP atau PDF.")
    st.divider()
    if st.session_state.processed and st.session_state.result_df is not None:
        n_c = len(st.session_state.dl_cache); tot_l = len(st.session_state.result_df)
        pct = int(n_c / max(tot_l, 1) * 100)
        st.markdown(f"**Cache:** {n_c}/{tot_l} ({pct}%)")
        st.progress(min(pct / 100, 1.0))
        if n_c > 0:
            sz = sum(len(v) for v in st.session_state.dl_cache.values())
            st.caption(f"Ukuran: {sz/1024/1024:.1f} MB")
        if st.button("🗑️ Bersihkan Cache", use_container_width=True):
            st.session_state.dl_cache = {}; st.rerun()
    st.caption("Made with ❤️ · Streamlit")

# ══════════════════════════════════════════════════════════════════════════════
# TOPBAR
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="topbar">
  <div class="topbar-brand">
    <div class="topbar-icon">🚛</div>
    <div>
      <div class="topbar-name">SuratJalan Bulk Downloader</div>
      <div class="topbar-ver">v2.2 · Logistics Tool</div>
    </div>
  </div>
  <div class="topbar-chips">
    <span class="topchip tc-orange">⚡ 8 Thread Paralel</span>
    <span class="topchip tc-green">✅ Auto-Hyperlink</span>
    <span class="topchip tc-blue">🔍 Fuzzy NOPOL</span>
  </div>
</div>
""", unsafe_allow_html=True)

# ── HERO ──
st.markdown("""
<div class="hero">
  <div class="hero-eyebrow">🚛 Logistics · Automation · Document</div>
  <div class="hero-title">Download Surat Jalan <span>in Bulk</span></div>
  <div class="hero-desc">Cocokkan NOPOL + Kuantum dari dua file, unduh ribuan dokumen sekaligus. Mendukung link hyperlink Excel, ZIP, PDF gabungan, deteksi duplikat &amp; saran NOPOL mirip dengan bulk download.</div>
  <div class="hero-stats">
    <div class="hero-stat"><div class="hero-stat-val">8×</div><div class="hero-stat-lbl">Thread Paralel</div></div>
    <div class="hero-stat"><div class="hero-stat-val">∞</div><div class="hero-stat-lbl">File per Batch</div></div>
    <div class="hero-stat"><div class="hero-stat-val">ZIP+PDF</div><div class="hero-stat-lbl">Output Format</div></div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── STEP PILLS ──
def render_steps(step):
    steps = [("1","Upload File","File 1 & 2"), ("2","Proses","Matching"), ("3","Download","ZIP / PDF")]
    parts = []
    for i, (n, lbl, sub) in enumerate(steps):
        idx = i + 1
        if idx < step: status, icon = "done", "✓"
        elif idx == step: status, icon = "active", n
        else: status, icon = "idle", n
        parts.append(f'<div class="step-pill {status}"><span class="step-num">{icon}</span><span>{lbl} <span style="opacity:.5;font-size:.67em">· {sub}</span></span></div>')
        if i < len(steps) - 1: parts.append('<span class="step-arrow">›</span>')
    st.markdown(f'<div class="steps">{"".join(parts)}</div>', unsafe_allow_html=True)

render_steps(3 if st.session_state.processed else 1)

# ══════════════════════════════════════════════════════════════════════════════
# UPLOAD SECTION
# ══════════════════════════════════════════════════════════════════════════════
st.markdown('<div class="sec-hdr"><span class="sec-badge">01</span><span class="sec-title">Upload File</span><span class="sec-line"></span></div>', unsafe_allow_html=True)

col_f1, col_f2 = st.columns(2, gap="medium")

with col_f1:
    file1 = st.session_state.get('f1_widget')
    ready1 = "ready" if file1 else ""
    st.markdown(f'''<div class="ucard {ready1}">
      <div class="ucard-head">
        <div class="ucard-icon ucard-icon-a">📋</div>
        <div>
          <div class="ucard-ttl">File 1 — Daftar Target</div>
          <div class="ucard-sub">Data yang ingin dicocokkan &amp; didownload</div>
        </div>
      </div>
      <div class="tag-row">
        <span class="tag-r">NOPOL *</span>
        <span class="tag-r">KUANTUM *</span>
      </div>
    </div>''', unsafe_allow_html=True)
    file1 = st.file_uploader("File 1", type=['csv','xlsx','xls'], key='f1_widget', label_visibility='collapsed')
    if file1:
        st.markdown(f'<div class="file-pill">✅ &nbsp;<b>{file1.name}</b> — {file1.size/1024:.1f} KB</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="alert a-info" style="margin-top:6px"><span class="alert-ico">💡</span><div style="font-size:.79rem">Upload <code>.xlsx</code>, <code>.xls</code>, atau <code>.csv</code> berisi NOPOL &amp; Kuantum</div></div>', unsafe_allow_html=True)

with col_f2:
    file2_raw = st.session_state.get('f2')
    ready2 = "ready" if file2_raw else ""
    st.markdown(f'''<div class="ucard {ready2}">
      <div class="ucard-head">
        <div class="ucard-icon ucard-icon-b">🗄️</div>
        <div>
          <div class="ucard-ttl">File 2 — Database Surat Jalan</div>
          <div class="ucard-sub">Berisi link Google Drive (URL atau hyperlink)</div>
        </div>
      </div>
      <div class="tag-row">
        <span class="tag-r">NOPOL *</span>
        <span class="tag-r">KUANTUM *</span>
        <span class="tag-r">Link GDrive *</span>
        <span class="tag-o">hyperlink ✓</span>
      </div>
    </div>''', unsafe_allow_html=True)
    file2 = st.file_uploader("File 2", type=['csv','xlsx','xls'], key='f2', label_visibility='collapsed')
    if file2:
        st.markdown(f'<div class="file-pill">✅ &nbsp;<b>{file2.name}</b> — {file2.size/1024:.1f} KB</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="alert a-info" style="margin-top:6px"><span class="alert-ico">💡</span><div style="font-size:.79rem">Upload <code>.xlsx</code>, <code>.xls</code>, atau <code>.csv</code> berisi link GDrive</div></div>', unsafe_allow_html=True)

# ── Process button ──
st.markdown("")
both = (file1 is not None) and (file2 is not None)
bc1, bc2 = st.columns([2, 10])
with bc1:
    process = st.button('⚙️ Proses & Cocokkan', use_container_width=True, disabled=not both, type="primary")
with bc2:
    if not both:
        missing_parts = []
        if file1 is None: missing_parts.append("File 1")
        if file2 is None: missing_parts.append("File 2")
        st.markdown(f'<div style="padding:10px 0;font-size:.8rem;color:#9c9790">Upload <b style="color:#e55a00">{" &amp; ".join(missing_parts)}</b> untuk melanjutkan</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div style="padding:10px 0;font-size:.8rem;color:#157a3c;font-weight:600">✅ Kedua file siap — klik tombol untuk memproses</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# PROSES
# ══════════════════════════════════════════════════════════════════════════════
if process:
    with st.spinner('🔄 Memproses dan mencocokkan data…'):
        if file2.name.lower().endswith(('.xlsx', '.xls')):
            raw_df2 = read_xlsx_with_hyperlinks(file2)
        else:
            raw_df2 = read_file(file2)
        df1 = load_file1(read_file(file1))
        df2 = load_file2(raw_df2)
        if df1.empty or df2.empty: st.stop()
        dup_df = detect_duplicates_f1(df1); st.session_state.dup_df = dup_df
        result_all = match_files(df1, df2)
        found = result_all[result_all['surat_jalan'].notna() & result_all['surat_jalan'].str.startswith('http', na=False)].copy().reset_index(drop=True)
        matched_f1_idx = set(found['_f1_idx'].tolist())
        missing_rows = df1[~df1.index.isin(matched_f1_idx)].copy()
        missing = missing_rows.drop_duplicates(subset=['nopol','kuantum']).reset_index(drop=True)
        missing_detail = build_missing_with_suggestions(missing, df1, df2)
        st.session_state.missing_detail = missing_detail
        diff_rows, miss_rows = [], []
        for item in missing_detail:
            if item['kategori'] == 'kuantum_beda':
                diff_rows.append({'NOPOL': item['nopol'], 'Kuantum File 1': item['kuantum'], 'Kuantum di File 2': item['info'], 'Status': '⚠️ Kuantum tidak cocok'})
            else:
                miss_rows.append({'NOPOL': item['nopol'], 'Kuantum File 1': item['kuantum'], 'Status': '❌ NOPOL tidak ada di File 2', 'Saran NOPOL (Kuantum Cocok)': (', '.join([f"{s['nopol_f2']} ({s['similarity']}%)" for s in item['saran']]) if item['saran'] else '-')})
        st.session_state.result_df = found; st.session_state.missing_df = missing
        st.session_state.nopol_diff_df = pd.DataFrame(diff_rows); st.session_state.nopol_miss_df = pd.DataFrame(miss_rows)
        st.session_state.df2_debug = df2; st.session_state.df1_debug = df1
        st.session_state.active_preview = None; st.session_state.dl_cache = {}
        st.session_state.saran_preview = {}; st.session_state.dup_prev_active = {}
        st.session_state.saran_selected = {}
        st.session_state.processed = True
    n_dup_groups = (len(dup_df[['nopol','kuantum']].drop_duplicates()) if not dup_df.empty else 0)
    n_mirip = sum(1 for x in missing_detail if x['kategori'] == 'nopol_mirip')
    st.balloons()
    st.success(f"✅ {len(found)} link ditemukan dari {len(df1)} data · {len(missing)} tidak match · {n_dup_groups} duplikat · {n_mirip} saran mirip")
    st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# RESULTS
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.result_df is not None:
    found = st.session_state.result_df; missing = st.session_state.missing_df
    nopol_diff = st.session_state.nopol_diff_df; nopol_miss = st.session_state.nopol_miss_df
    df2_all = st.session_state.df2_debug; df1_all = st.session_state.df1_debug
    dup_df = st.session_state.dup_df if st.session_state.dup_df is not None else pd.DataFrame()
    missing_detail = st.session_state.missing_detail or []
    n_match = len(found); n_diff_k = len(nopol_diff) if nopol_diff is not None else 0
    n_miss_nopol = len(nopol_miss) if nopol_miss is not None else 0; n_all_miss = len(missing)
    n_dup_groups = (len(dup_df[['nopol','kuantum']].drop_duplicates()) if not dup_df.empty else 0)
    n_dup_rows = len(dup_df) if not dup_df.empty else 0
    n_nopol_mirip = sum(1 for x in missing_detail if x['kategori'] == 'nopol_mirip')
    match_rate = int(n_match / max(len(df1_all), 1) * 100) if df1_all is not None else 0

    st.markdown('<div class="sec-hdr"><span class="sec-badge">02</span><span class="sec-title">Ringkasan Hasil</span><span class="sec-line"></span></div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="stat-grid">
      <div class="stat-card"><div class="stat-stripe" style="background:linear-gradient(90deg,#157a3c,#22c55e)"></div><div class="stat-icon-bg">✅</div><div class="stat-lbl">Link Ditemukan</div><div class="stat-val" style="color:#157a3c">{n_match}</div><div class="stat-sub" style="color:#157a3c">{match_rate}% match rate</div></div>
      <div class="stat-card"><div class="stat-stripe" style="background:linear-gradient(90deg,#6d28d9,#a78bfa)"></div><div class="stat-icon-bg">🔁</div><div class="stat-lbl">Duplikat File 1</div><div class="stat-val" style="color:#6d28d9">{n_dup_groups}</div><div class="stat-sub" style="color:#6d28d9">{n_dup_rows} baris total</div></div>
      <div class="stat-card"><div class="stat-stripe" style="background:linear-gradient(90deg,#b45309,#fbbf24)"></div><div class="stat-icon-bg">⚠️</div><div class="stat-lbl">Kuantum Beda</div><div class="stat-val" style="color:#b45309">{n_diff_k}</div><div class="stat-sub" style="color:#b45309">NOPOL ada, qty ≠</div></div>
      <div class="stat-card"><div class="stat-stripe" style="background:linear-gradient(90deg,#c41c1c,#f87171)"></div><div class="stat-icon-bg">❌</div><div class="stat-lbl">NOPOL Tidak Ada</div><div class="stat-val" style="color:#c41c1c">{n_miss_nopol}</div><div class="stat-sub" style="color:#c41c1c">{n_nopol_mirip} ada saran</div></div>
      <div class="stat-card"><div class="stat-stripe" style="background:linear-gradient(90deg,#e55a00,#fb923c)"></div><div class="stat-icon-bg">🔴</div><div class="stat-lbl">Total Tidak Match</div><div class="stat-val" style="color:#e55a00">{n_all_miss}</div><div class="stat-sub" style="color:#e55a00">dari {len(df1_all)} data</div></div>
    </div>
    """, unsafe_allow_html=True)

    if n_dup_groups > 0:
        st.markdown(f'<div class="alert a-violet"><span class="alert-ico">🔁</span><div><b>{n_dup_groups} kombinasi duplikat</b> ({n_dup_rows} baris) di File 1. Nama file dibedakan <code>_DUPLIKAT1</code>, <code>_DUPLIKAT2</code>, dst.</div></div>', unsafe_allow_html=True)
    if n_nopol_mirip > 0:
        st.markdown(f'<div class="alert a-sky"><span class="alert-ico">🔍</span><div><b>{n_nopol_mirip} data</b> punya saran NOPOL mirip. Buka tab <b>NOPOL Tidak Ada</b> → pilih saran &amp; download bulk.</div></div>', unsafe_allow_html=True)
    if n_match == 0:
        st.markdown('<div class="alert a-error"><span class="alert-ico">⚠️</span><div><b>Tidak ada data yang match.</b> Pastikan format NOPOL &amp; KUANTUM konsisten di kedua file.</div></div>', unsafe_allow_html=True)

    st.markdown('<div class="sec-hdr"><span class="sec-badge">03</span><span class="sec-title">Detail & Download</span><span class="sec-line"></span></div>', unsafe_allow_html=True)

    tab1, tab_dup, tab2, tab3, tab4 = st.tabs([
        f"✅ Match ({n_match})",
        f"🔁 Duplikat ({n_dup_groups}g · {n_dup_rows}b)",
        f"⚠️ Kuantum Beda ({n_diff_k})",
        f"❌ NOPOL Tidak Ada ({n_miss_nopol})",
        f"🔴 Semua Tidak Match ({n_all_miss})",
    ])

    # ── TAB 1: MATCH ─────────────────────────────────────────────────────
    with tab1:
        if len(found) == 0:
            st.markdown('<div class="empty-state"><div class="ico">🔍</div><h3>Tidak Ada Data yang Match</h3><p>Tidak ada kombinasi NOPOL + Kuantum yang cocok antara File 1 dan File 2.</p></div>', unsafe_allow_html=True)
        else:
            sc1, sc2 = st.columns([3, 9])
            with sc1:
                search = st.text_input('Filter', placeholder='🔍  Ketik NOPOL…', label_visibility='collapsed', key='search_found')
            disp = found.copy()
            if search.strip():
                disp = disp[disp['nopol'].str.contains(re.escape(norm_nopol(search.strip())), na=False, case=False)].reset_index(drop=True)
            with sc2:
                st.markdown(f'<div style="padding:10px 4px;font-size:.78rem;color:#9c9790">Menampilkan <b style="color:#1a1917">{len(disp)}</b> dari <b style="color:#1a1917">{len(found)}</b> surat jalan</div>', unsafe_allow_html=True)

            ac1, ac2, ac3, ac4 = st.columns([2, 2, 2, 6])
            with ac1: do_zip = st.button('📦 Download ZIP', use_container_width=True, key='btn_zip')
            with ac2: do_merge = st.button('📄 Gabung PDF', use_container_width=True, key='btn_merge')
            with ac3: do_preload = st.button('⚡ Pre-load', use_container_width=True, key='btn_preload')
            with ac4:
                n_c = sum(1 for _, r in disp.iterrows() if r['surat_jalan'] in st.session_state.dl_cache)
                pct_c = int(n_c / max(len(disp), 1) * 100)
                color = "#157a3c" if pct_c == 100 else ("#b45309" if pct_c > 0 else "#9c9790")
                st.markdown(f'<div class="cache-info">🗃️ Cache: <b style="color:{color}">{n_c}/{len(disp)}</b> ({pct_c}%) {"· ✅ Semua siap" if pct_c==100 else "· Gunakan ⚡ Pre-load untuk mempercepat"}</div>', unsafe_allow_html=True)

            if do_preload and len(disp) > 0:
                needed = [r['surat_jalan'] for _, r in disp.iterrows() if r['surat_jalan'] not in st.session_state.dl_cache]
                if not needed:
                    st.success('✅ Semua sudah di cache!')
                else:
                    snap = dict(st.session_state.dl_cache)
                    tasks_pre = [{'idx': i, 'nopol': r['nopol'], 'kuantum': int(r['kuantum']), 'link': r['surat_jalan'], 'dup_label': ''} for i, r in disp.iterrows() if r['surat_jalan'] not in snap]
                    pb = st.progress(0); stxt = st.empty(); ok_n = fail_n = done_n = 0; nc = {}
                    with ThreadPoolExecutor(max_workers=8) as ex:
                        futs = {ex.submit(_worker, t, snap): t for t in tasks_pre}
                        for fut in as_completed(futs):
                            res = fut.result()
                            if res['content']: nc[res['link']] = res['content']; ok_n += 1
                            else: fail_n += 1
                            done_n += 1; pb.progress(done_n / len(tasks_pre))
                            stxt.markdown(f"Pre-load **{done_n}/{len(tasks_pre)}** — ✅ {ok_n} · ❌ {fail_n}")
                    st.session_state.dl_cache.update(nc); stxt.success(f'✅ Selesai: {ok_n} berhasil, {fail_n} gagal')

            if do_zip and len(disp) > 0:
                rows_dl = [{'idx': i, 'nopol': r['nopol'], 'kuantum': int(r['kuantum']), 'link': r['surat_jalan'], 'dup_label': ''} for i, r in disp.iterrows()]
                ok_files, fl, nc = run_bulk_download(rows_dl, 'ZIP'); st.session_state.dl_cache.update(nc)
                if ok_files:
                    zd = make_zip(ok_files)
                    st.markdown(f'<div class="alert a-success"><span class="alert-ico">📦</span><div><b>{len(ok_files)} file siap</b> · ZIP {len(zd)//1024:,} KB</div></div>', unsafe_allow_html=True)
                    st.download_button(f'💾 Simpan ZIP ({len(ok_files)} file · {len(zd)//1024:,} KB)', zd, 'surat_jalan.zip', 'application/zip', key='dl_zip_r')
                if fl:
                    with st.expander(f'⚠️ {len(fl)} gagal'): [st.markdown(f'`{f}`') for f in fl]

            if do_merge and len(disp) > 0:
                rows_dl = [{'idx': i, 'nopol': r['nopol'], 'kuantum': int(r['kuantum']), 'link': r['surat_jalan'], 'dup_label': ''} for i, r in disp.iterrows()]
                ok_files, fail_dl, nc = run_bulk_download(rows_dl, 'PDF'); st.session_state.dl_cache.update(nc)
                if ok_files:
                    with st.spinner('📄 Menggabungkan…'):
                        ordered = [nc.get(r['surat_jalan']) or st.session_state.dl_cache.get(r['surat_jalan']) for _, r in disp.iterrows()]
                        ordered = [x for x in ordered if x]; merged = merge_pdfs(ordered)
                    if merged:
                        st.markdown(f'<div class="alert a-success"><span class="alert-ico">📄</span><div><b>{len(ordered)} file digabung</b> · PDF {len(merged)//1024:,} KB</div></div>', unsafe_allow_html=True)
                        st.download_button(f'💾 Simpan PDF ({len(ordered)} hal · {len(merged)//1024:,} KB)', merged, 'surat_jalan_gabungan.pdf', 'application/pdf', key='dl_merged')
                        if fail_dl:
                            with st.expander(f'⚠️ {len(fail_dl)} tidak ikut'): [st.markdown(f'`{f}`') for f in fail_dl]
                    else:
                        st.error('❌ Gagal PDF. Pastikan `pypdf`, `reportlab`, `Pillow` terinstall.')

            st.markdown('<hr class="row-div">', unsafe_allow_html=True)
            hcols = st.columns([0.5, 2.5, 1.3, 0.7, 1.8, 0.7, 2])
            for col, lbl in zip(hcols, ['#','NOPOL','Kuantum','Link #','Drive','👁','⬇️']):
                col.markdown(f'<span class="col-hdr">{lbl}</span>', unsafe_allow_html=True)
            st.markdown('<hr class="row-div">', unsafe_allow_html=True)

            for i, row in disp.iterrows():
                nopol = row['nopol']; kuantum = int(row['kuantum']); link = row['surat_jalan']
                link_no = int(row.get('_link_no', 1)); fid = extract_fid(link)
                is_dup = not dup_df.empty and ((dup_df['nopol'] == nopol) & (dup_df['kuantum'] == kuantum)).any()
                cols = st.columns([0.5, 2.5, 1.3, 0.7, 1.8, 0.7, 2])
                cols[0].markdown(f'<span style="font-size:.73rem;color:#9c9790">#{i+1}</span>', unsafe_allow_html=True)
                dup_badge = ' <span style="background:var(--violet-lt);color:#6d28d9;font-size:.6rem;padding:1px 5px;border-radius:4px;font-weight:700">DUP</span>' if is_dup else ''
                cols[1].markdown(f'<span class="np {"np-dup" if is_dup else ""}">{nopol}</span>{dup_badge}', unsafe_allow_html=True)
                cols[2].markdown(f'<span style="font-family:var(--mono);font-size:.83rem">{kuantum:,}</span>', unsafe_allow_html=True)
                cols[3].markdown(f'<span style="font-size:.7rem;color:#9c9790">#{link_no}</span>', unsafe_allow_html=True)
                if fid: cols[4].markdown(f'[🔗 Buka](https://drive.google.com/file/d/{fid}/view)')
                else: cols[4].markdown(f'[🔗 Buka]({link})')
                with cols[5]:
                    if st.button('👁️✕' if st.session_state.active_preview == i else '👁️', key=f'v_{i}'):
                        st.session_state.active_preview = (None if st.session_state.active_preview == i else i); st.rerun()
                with cols[6]:
                    dup_label = f'_DUPLIKAT{link_no}' if link_no > 1 else ''
                    cached = st.session_state.dl_cache.get(link)
                    if cached:
                        ext = infer_extension(cached)
                        fname = make_safe_filename(nopol, kuantum, i, ext, total=len(disp), dup_label=dup_label)
                        st.download_button(f'⬇️ .{ext.upper()}', cached, fname, 'application/pdf' if ext=='pdf' else f'image/{ext}', key=f'd_{i}')
                    else:
                        if st.button('⬇️ Unduh', key=f'db_{i}'):
                            with st.spinner(f'Mengunduh {nopol}…'): ct = download_file(link)
                            if ct:
                                st.session_state.dl_cache[link] = ct; ext = infer_extension(ct)
                                fname = make_safe_filename(nopol, kuantum, i, ext, total=len(disp), dup_label=dup_label)
                                st.download_button(f'💾 .{ext.upper()}', ct, fname, 'application/pdf' if ext=='pdf' else f'image/{ext}', key=f'ds_{i}')
                            else: st.error('❌ Gagal — file private/expired.')
                if st.session_state.active_preview == i:
                    purl = to_preview(link)
                    if purl:
                        import streamlit.components.v1 as components
                        st.markdown(f'<div class="alert a-sky" style="margin-top:8px"><span class="alert-ico">👁️</span><div>Preview — <b>{nopol}</b> · {kuantum:,} <a href="{purl}" target="_blank" style="margin-left:8px;font-size:.76rem">↗ Tab baru</a></div></div>', unsafe_allow_html=True)
                        components.html(f'<iframe src="{purl}" width="100%" height="680" style="border:1.5px solid #bae6fd;border-radius:11px;background:#fff" allow="autoplay"></iframe>', height=700)
                    else: st.error('Link preview tidak valid.')
                if i < len(disp) - 1: st.markdown('<hr class="row-div">', unsafe_allow_html=True)

    # ── TAB DUP ──────────────────────────────────────────────────────────
    with tab_dup:
        if dup_df.empty:
            st.markdown('<div class="empty-state"><div class="ico">🎉</div><h3>Tidak Ada Duplikat!</h3><p>Semua kombinasi NOPOL + Kuantum di File 1 adalah unik.</p></div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="alert a-violet"><span class="alert-ico">🔁</span><div><b>{n_dup_groups} kombinasi duplikat</b> ({n_dup_rows} baris). Dibedakan <code>_DUPLIKAT1</code>, <code>_DUPLIKAT2</code>, dst.</div></div>', unsafe_allow_html=True)
            sd_dup = st.text_input('', placeholder='🔍  Filter NOPOL duplikat…', label_visibility='collapsed', key='sd_dup')
            dza, dzb, _ = st.columns([2, 2, 8])
            with dza: do_zip_dup = st.button('📦 ZIP Duplikat', use_container_width=True, key='btn_zip_dup')
            with dzb: do_merge_dup = st.button('📄 PDF Duplikat', use_container_width=True, key='btn_merge_dup')

            all_dup_rows = []
            for (nopol, kuantum), grp in dup_df.groupby(['nopol', 'kuantum']):
                lf = found[(found['nopol']==nopol)&(found['kuantum']==kuantum)]['surat_jalan'].tolist()
                for i_row, (_, drow) in enumerate(grp.iterrows(), start=1):
                    bk = int(drow['baris_ke'])
                    lnk = None if not lf else lf[i_row-1] if i_row<=len(lf) else lf[0]
                    all_dup_rows.append({'nopol': nopol, 'kuantum': int(kuantum), 'baris_ke': bk, 'dup_label': f'_DUPLIKAT{bk}', 'link': lnk})

            if do_zip_dup:
                vr = [r for r in all_dup_rows if r['link']]
                if not vr: st.warning('⚠️ Tidak ada link.')
                else:
                    rows_dl = [{'idx': i, 'nopol': r['nopol'], 'kuantum': r['kuantum'], 'link': r['link'], 'dup_label': r['dup_label']} for i, r in enumerate(vr)]
                    ok_files, fl, nc = run_bulk_download(rows_dl, 'Duplikat ZIP'); st.session_state.dl_cache.update(nc)
                    if ok_files: st.download_button(f'💾 ZIP Duplikat ({len(ok_files)} file)', make_zip(ok_files), 'duplikat.zip', 'application/zip', key='dl_zip_dup_r')
                    if fl:
                        with st.expander(f'❌ {len(fl)} gagal'): [st.markdown(f'`{f}`') for f in fl]

            if do_merge_dup:
                vr = [r for r in all_dup_rows if r['link']]
                if not vr: st.warning('⚠️ Tidak ada link.')
                else:
                    rows_dl = [{'idx': i, 'nopol': r['nopol'], 'kuantum': r['kuantum'], 'link': r['link'], 'dup_label': r['dup_label']} for i, r in enumerate(vr)]
                    ok_files, fl, nc = run_bulk_download(rows_dl, 'Duplikat PDF'); st.session_state.dl_cache.update(nc)
                    if ok_files:
                        with st.spinner('Menggabungkan…'):
                            ordered = [nc.get(r['link']) or st.session_state.dl_cache.get(r['link']) for r in vr]; ordered = [x for x in ordered if x]; merged = merge_pdfs(ordered)
                        if merged: st.success(f'✅ {len(ordered)} file digabung · {len(merged)//1024:,} KB'); st.download_button('💾 PDF Duplikat', merged, 'duplikat.pdf', 'application/pdf', key='dl_merge_dup_r')
                        else: st.error('❌ Gagal PDF.')
                    if fl:
                        with st.expander(f'⚠️ {len(fl)} gagal'): [st.markdown(f'`{f}`') for f in fl]

            st.markdown('<hr class="row-div">', unsafe_allow_html=True)
            row_counter = 0
            for (nopol, kuantum), grp in dup_df.groupby(['nopol', 'kuantum']):
                if sd_dup.strip() and norm_nopol(sd_dup.strip()) not in nopol: continue
                lf = found[(found['nopol']==nopol)&(found['kuantum']==kuantum)]['surat_jalan'].tolist()
                st.markdown(f'<div class="dup-group"><div class="dup-group-header"><span class="np np-dup">{nopol}</span><span style="color:#4a4641;font-size:.8rem;font-family:var(--mono)">{int(kuantum):,}</span><span style="background:#fee2e2;color:#991b1b;border-radius:5px;padding:2px 8px;font-size:.7rem;font-weight:700">Muncul {len(grp)}×</span><span style="background:#eff6ff;color:#1d4ed8;border-radius:5px;padding:2px 8px;font-size:.7rem;font-weight:700">{len(lf)} link di File 2</span></div>', unsafe_allow_html=True)
                dh = st.columns([0.6, 1.8, 2.2, 2, 0.8, 1.8])
                for col, lbl in zip(dh, ['Ke-','NOPOL','Label','Drive','👁','⬇️']): col.markdown(f'<span class="col-hdr">{lbl}</span>', unsafe_allow_html=True)
                for i_row, (_, drow) in enumerate(grp.iterrows(), start=1):
                    bk = int(drow['baris_ke']); dlbl = f'_DUPLIKAT{bk}'
                    lnk = None if not lf else lf[i_row-1] if i_row<=len(lf) else lf[0]
                    fid = extract_fid(lnk) if lnk else None
                    uid = f'dup_{re.sub(r"[^a-zA-Z0-9]","_",nopol)}_{kuantum}_{bk}'
                    cols = st.columns([0.6, 1.8, 2.2, 2, 0.8, 1.8])
                    cols[0].markdown(f'<b style="color:#6d28d9">{bk}</b>', unsafe_allow_html=True)
                    cols[1].markdown(f'<span class="np np-dup">{nopol}</span>', unsafe_allow_html=True)
                    cols[2].markdown(f'<code style="background:var(--violet-lt);color:#6d28d9;padding:2px 7px;border-radius:4px;font-size:.76rem">{dlbl}</code>', unsafe_allow_html=True)
                    if lnk and fid: cols[3].markdown(f'[🔗 Drive](https://drive.google.com/file/d/{fid}/view)')
                    elif lnk: cols[3].markdown(f'[🔗 Buka]({lnk})')
                    else: cols[3].markdown('<span style="color:#9c9790">— Tidak ada</span>', unsafe_allow_html=True)
                    with cols[4]:
                        if lnk:
                            pgk = f'grp_{re.sub(r"[^a-zA-Z0-9]","_",nopol)}_{kuantum}'
                            isap = (st.session_state.dup_prev_active.get(pgk) == bk)
                            if st.button('👁️✕' if isap else '👁️', key=f'dprev_{uid}'):
                                st.session_state.dup_prev_active[pgk] = (None if isap else bk); st.rerun()
                    with cols[5]:
                        if lnk:
                            cached = st.session_state.dl_cache.get(lnk)
                            if cached:
                                ext = infer_extension(cached); fname = make_safe_filename(nopol, kuantum, row_counter, ext, total=n_dup_rows, dup_label=dlbl)
                                st.download_button(f'⬇️ .{ext.upper()}', cached, fname, 'application/pdf' if ext=='pdf' else f'image/{ext}', key=f'ddl_{uid}')
                            else:
                                if st.button('⬇️ Unduh', key=f'ddlb_{uid}'):
                                    with st.spinner('Mengunduh…'): ct = download_file(lnk)
                                    if ct:
                                        st.session_state.dl_cache[lnk] = ct; ext = infer_extension(ct); fname = make_safe_filename(nopol, kuantum, row_counter, ext, total=n_dup_rows, dup_label=dlbl)
                                        st.download_button(f'💾 .{ext.upper()}', ct, fname, 'application/pdf' if ext=='pdf' else f'image/{ext}', key=f'ddls_{uid}')
                                    else: st.error('❌ Gagal.')
                        else: st.markdown('`—`')
                    pgk = f'grp_{re.sub(r"[^a-zA-Z0-9]","_",nopol)}_{kuantum}'
                    if lnk and st.session_state.dup_prev_active.get(pgk) == bk:
                        purl = to_preview(lnk)
                        if purl:
                            import streamlit.components.v1 as components
                            st.markdown(f'<div class="alert a-sky" style="margin-top:8px"><span class="alert-ico">👁️</span><div>Preview — <b>{nopol}</b> · <code style="color:#6d28d9">{dlbl}</code> · <a href="{purl}" target="_blank" style="font-size:.76rem">↗ Tab baru</a></div></div>', unsafe_allow_html=True)
                            components.html(f'<iframe src="{purl}" width="100%" height="680" style="border:1.5px solid #ddd6fe;border-radius:11px;background:#fff" allow="autoplay"></iframe>', height=700)
                        else: st.error('Link preview tidak valid.')
                    row_counter += 1
                st.markdown('</div>', unsafe_allow_html=True); st.markdown("")

            st.markdown('<hr class="row-div">', unsafe_allow_html=True)
            exp_dup = dup_df[['nopol','kuantum','baris_ke','jumlah_duplikat']].copy()
            exp_dup.columns = ['NOPOL','Kuantum','Baris ke-','Total Duplikat']
            exp_dup['Label File'] = exp_dup['Baris ke-'].apply(lambda x: f'_DUPLIKAT{int(x)}')
            ec, _ = st.columns([2, 10])
            with ec: st.download_button('📥 Export CSV', exp_dup.to_csv(index=False).encode('utf-8'), 'duplikat.csv', 'text/csv', key='dl_dup')
            st.dataframe(exp_dup, use_container_width=True, hide_index=True)

    # ── TAB 2: KUANTUM BEDA ───────────────────────────────────────────────
    with tab2:
        if nopol_diff is None or nopol_diff.empty:
            st.markdown('<div class="empty-state"><div class="ico">🎉</div><h3>Semua Kuantum Cocok!</h3><p>Tidak ada NOPOL dengan kuantum berbeda di File 2.</p></div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="alert a-warn"><span class="alert-ico">⚠️</span><div><b>{len(nopol_diff)} data</b> — NOPOL ditemukan di File 2 tapi kuantumnya tidak cocok.</div></div>', unsafe_allow_html=True)
            sc, _ = st.columns([3, 9])
            with sc: sd = st.text_input('', placeholder='🔍 Filter NOPOL…', label_visibility='collapsed', key='sd')
            dd = nopol_diff.copy()
            if sd.strip(): dd = dd[dd['NOPOL'].str.contains(re.escape(norm_nopol(sd)), na=False, case=False)].reset_index(drop=True)
            st.dataframe(dd, use_container_width=True, hide_index=True)
            ca, _ = st.columns([2, 10])
            with ca: st.download_button('📥 Export CSV', dd.to_csv(index=False).encode('utf-8'), 'kuantum_beda.csv', 'text/csv', key='dl_a')

    # ══════════════════════════════════════════════════════════════════════
    # ── TAB 3: NOPOL TIDAK ADA — with CHECKBOX + BULK DOWNLOAD ───────────
    # ══════════════════════════════════════════════════════════════════════
    with tab3:
        miss_items = [x for x in missing_detail if x['kategori'] in ('nopol_mirip', 'tidak_ada')]
        if not miss_items:
            st.markdown('<div class="empty-state"><div class="ico">🎉</div><h3>Semua NOPOL Ditemukan!</h3><p>Tidak ada NOPOL dari File 1 yang hilang di File 2.</p></div>', unsafe_allow_html=True)
        else:
            n_mirip_t3 = sum(1 for x in miss_items if x['kategori'] == 'nopol_mirip')
            n_tdk_ada_t3 = sum(1 for x in miss_items if x['kategori'] == 'tidak_ada')
            st.markdown(f'<div class="alert a-error"><span class="alert-ico">❌</span><div><b>{len(miss_items)} NOPOL tidak ada di File 2.</b> 🔍 <b>{n_mirip_t3}</b> punya saran mirip · 🚫 <b>{n_tdk_ada_t3}</b> tanpa saran</div></div>', unsafe_allow_html=True)

            # ── CONTROLS ROW ──────────────────────────────────────────────
            fc1, fc2, fc3 = st.columns([3, 4, 3])
            with fc1:
                sm = st.text_input('', placeholder='🔍 Filter NOPOL…', label_visibility='collapsed', key='sm_t3')
            with fc2:
                min_sim = st.slider('Min Kemiripan (%)', 30, 95, 80, 5, key='sim_slider_t3')
            with fc3:
                st.markdown('<div style="height:6px"></div>', unsafe_allow_html=True)
                auto_select = st.button(f'✨ Pilih Otomatis ≥{min_sim}%', use_container_width=True, key='auto_sel')

            # Build full list of selectable saran items
            all_saran_items = []
            for item_idx, item in enumerate(miss_items):
                for s_idx, saran in enumerate(item['saran']):
                    sk = f'chk_{item_idx}_{s_idx}'
                    all_saran_items.append({
                        'key': sk,
                        'item_idx': item_idx,
                        's_idx': s_idx,
                        'nopol_f1': item['nopol'],
                        'kuantum': item['kuantum'],
                        'nopol_f2': saran['nopol_f2'],
                        'similarity': saran['similarity'],
                        'surat_jalan': saran['surat_jalan'],
                    })

            # Auto-select button logic
            if auto_select:
                for si in all_saran_items:
                    if si['similarity'] >= min_sim:
                        st.session_state.saran_selected[si['key']] = True
                    # don't deselect manually chosen ones below threshold
                st.rerun()

            # Quick select / deselect all buttons
            sel_col1, sel_col2, sel_col3, sel_col4 = st.columns([2, 2, 2, 6])
            with sel_col1:
                if st.button('☑️ Pilih Semua', use_container_width=True, key='sel_all_t3'):
                    # Per NOPOL target: only select the first (best) suggestion
                    seen_item = set()
                    for si in all_saran_items:
                        if si['item_idx'] not in seen_item:
                            st.session_state.saran_selected[si['key']] = True
                            seen_item.add(si['item_idx'])
                        else:
                            # Deselect lower-ranked suggestions so only #1 is ticked
                            st.session_state.saran_selected[si['key']] = False
                    st.rerun()
            with sel_col2:
                if st.button('☐ Batalkan Semua', use_container_width=True, key='desel_all_t3'):
                    st.session_state.saran_selected = {}
                    st.rerun()

            # Count selected
            n_selected = sum(1 for si in all_saran_items if st.session_state.saran_selected.get(si['key'], False))

            with sel_col4:
                if n_selected > 0:
                    st.markdown(f'<div class="sel-controls"><span class="sel-label">Terpilih:</span><span class="sel-count">{n_selected} surat jalan</span><span style="font-size:.73rem;color:#9c9790;margin-left:4px">· gunakan tombol bulk download di bawah</span></div>', unsafe_allow_html=True)
                else:
                    st.markdown('<div style="padding:9px 14px;font-size:.76rem;color:#9c9790;background:var(--surface2);border:1.5px solid var(--border);border-radius:9px">Belum ada yang dipilih — centang saran di bawah</div>', unsafe_allow_html=True)

            # ── BULK ACTION BAR (shown when something is selected) ────────
            if n_selected > 0:
                selected_items = [si for si in all_saran_items if st.session_state.saran_selected.get(si['key'], False)]
                st.markdown(f'''<div class="bulk-bar">
                  <div>
                    <div class="bulk-bar-info">📦 <span>{n_selected}</span> surat jalan siap diunduh</div>
                    <div class="bulk-bar-sub">Klik ZIP untuk download semua sekaligus</div>
                  </div>
                </div>''', unsafe_allow_html=True)

                bba, bbb, bbc, _ = st.columns([2.5, 2.5, 2.5, 5])
                with bba:
                    do_bulk_zip_saran = st.button(f'📦 ZIP ({n_selected} file)', use_container_width=True, key='bulk_zip_saran', type='primary')
                with bbb:
                    do_bulk_merge_saran = st.button(f'📄 Gabung PDF ({n_selected})', use_container_width=True, key='bulk_merge_saran')
                with bbc:
                    do_bulk_preload_saran = st.button(f'⚡ Pre-load ({n_selected})', use_container_width=True, key='bulk_preload_saran')

                if do_bulk_preload_saran:
                    snap = dict(st.session_state.dl_cache)
                    tasks_p = [{'idx': i, 'nopol': si['nopol_f2'], 'kuantum': si['kuantum'],
                                 'link': si['surat_jalan'], 'dup_label': ''} for i, si in enumerate(selected_items) if si['surat_jalan'] not in snap]
                    if not tasks_p:
                        st.success('✅ Semua sudah di cache!')
                    else:
                        pb2 = st.progress(0); stxt2 = st.empty(); ok2 = fail2 = done2 = 0; nc2 = {}
                        with ThreadPoolExecutor(max_workers=8) as ex:
                            futs2 = {ex.submit(_worker, t, snap): t for t in tasks_p}
                            for fut in as_completed(futs2):
                                res = fut.result()
                                if res['content']: nc2[res['link']] = res['content']; ok2 += 1
                                else: fail2 += 1
                                done2 += 1; pb2.progress(done2 / len(tasks_p))
                                stxt2.markdown(f"Pre-load **{done2}/{len(tasks_p)}** — ✅ {ok2} · ❌ {fail2}")
                        st.session_state.dl_cache.update(nc2); stxt2.success(f'✅ {ok2} berhasil, {fail2} gagal')

                if do_bulk_zip_saran:
                    rows_dl = [{'idx': i, 'nopol': si['nopol_f2'], 'kuantum': si['kuantum'],
                                 'link': si['surat_jalan'], 'dup_label': f'_SARAN{int(si["similarity"])}pct'} for i, si in enumerate(selected_items)]
                    ok_files, fl, nc = run_bulk_download(rows_dl, 'Saran ZIP'); st.session_state.dl_cache.update(nc)
                    if ok_files:
                        zd = make_zip(ok_files)
                        st.markdown(f'<div class="alert a-success"><span class="alert-ico">📦</span><div><b>{len(ok_files)} file siap</b> · ZIP {len(zd)//1024:,} KB</div></div>', unsafe_allow_html=True)
                        st.download_button(f'💾 Simpan ZIP Saran ({len(ok_files)} file · {len(zd)//1024:,} KB)', zd, 'saran_nopol.zip', 'application/zip', key='dl_saran_zip_r')
                    if fl:
                        with st.expander(f'⚠️ {len(fl)} gagal'): [st.markdown(f'`{f}`') for f in fl]

                if do_bulk_merge_saran:
                    rows_dl = [{'idx': i, 'nopol': si['nopol_f2'], 'kuantum': si['kuantum'],
                                 'link': si['surat_jalan'], 'dup_label': ''} for i, si in enumerate(selected_items)]
                    ok_files, fail_dl, nc = run_bulk_download(rows_dl, 'Saran PDF'); st.session_state.dl_cache.update(nc)
                    if ok_files:
                        with st.spinner('📄 Menggabungkan…'):
                            ordered_m = [nc.get(si['surat_jalan']) or st.session_state.dl_cache.get(si['surat_jalan']) for si in selected_items]
                            ordered_m = [x for x in ordered_m if x]; merged_s = merge_pdfs(ordered_m)
                        if merged_s:
                            st.markdown(f'<div class="alert a-success"><span class="alert-ico">📄</span><div><b>{len(ordered_m)} file digabung</b> · PDF {len(merged_s)//1024:,} KB</div></div>', unsafe_allow_html=True)
                            st.download_button(f'💾 Simpan PDF Saran ({len(ordered_m)} hal)', merged_s, 'saran_nopol_gabungan.pdf', 'application/pdf', key='dl_saran_merge_r')
                        else: st.error('❌ Gagal PDF.')
                    if fail_dl:
                        with st.expander(f'⚠️ {len(fail_dl)} tidak ikut'): [st.markdown(f'`{f}`') for f in fail_dl]

            st.markdown('<hr class="row-div">', unsafe_allow_html=True)

            # ── ITEM LIST with CHECKBOXES ─────────────────────────────────
            for item_idx, item in enumerate(miss_items):
                nopol = item['nopol']; kuantum = item['kuantum']
                if sm.strip() and norm_nopol(sm.strip()) not in nopol: continue

                saran_filtered = [s for s in item['saran'] if s['similarity'] >= min_sim]
                has_s = bool(saran_filtered)

                # Count how many saran of this item are selected
                n_item_selected = sum(
                    1 for s_idx in range(len(item['saran']))
                    if st.session_state.saran_selected.get(f'chk_{item_idx}_{s_idx}', False)
                )

                # Item header
                bc_item = '#dbeafe' if has_s else '#fee2e2'
                bl_item = '#2563eb' if has_s else '#c41c1c'
                bc2v_item = '#1d4ed8' if has_s else '#991b1b'
                selected_badge = f'<span style="background:var(--green-lt);color:var(--green);border:1px solid var(--green-md);border-radius:5px;padding:2px 8px;font-size:.7rem;font-weight:700">✓ {n_item_selected} dipilih</span>' if n_item_selected > 0 else ''
                st.markdown(f'''<div style="background:#fff;border:1.5px solid {bc_item};border-left:3px solid {bl_item};border-radius:11px;padding:12px 16px;margin:14px 0 4px;box-shadow:var(--shadow-sm)">
                  <div style="display:flex;align-items:center;gap:9px;flex-wrap:wrap">
                    <span class="np">{nopol}</span>
                    <span style="color:#4a4641;font-size:.78rem;font-family:var(--mono)">{kuantum:,}</span>
                    <span style="background:{bc_item};color:{bc2v_item};border-radius:5px;padding:2px 8px;font-size:.7rem;font-weight:700">{"🔍 " + str(len(saran_filtered)) + " saran (≥" + str(min_sim) + "%)" if has_s else "🚫 Tidak ada saran"}</span>
                    {selected_badge}
                  </div>
                </div>''', unsafe_allow_html=True)

                if saran_filtered:
                    # Column headers
                    sh_cols = st.columns([0.5, 0.4, 2.5, 1.4, 1.3, 1.8, 0.7, 1.6])
                    for col, lbl in zip(sh_cols, ['☑','#','NOPOL File 2','Kemiripan','Kuantum','Drive','👁','⬇️']):
                        col.markdown(f'<span class="col-hdr">{lbl}</span>', unsafe_allow_html=True)

                    for s_idx_full, saran in enumerate(item['saran']):
                        if saran['similarity'] < min_sim: continue
                        sk = f'chk_{item_idx}_{s_idx_full}'
                        sim = saran['similarity']
                        is_checked = st.session_state.saran_selected.get(sk, False)

                        row_cols = st.columns([0.5, 0.4, 2.5, 1.4, 1.3, 1.8, 0.7, 1.6])

                        # Checkbox
                        with row_cols[0]:
                            new_val = st.checkbox('', value=is_checked, key=f'cb_{sk}', label_visibility='collapsed')
                            if new_val != is_checked:
                                st.session_state.saran_selected[sk] = new_val; st.rerun()

                        row_cols[1].markdown(f'`{s_idx_full+1}`')
                        row_cols[2].markdown(f'<span class="np">{saran["nopol_f2"]}</span>', unsafe_allow_html=True)
                        row_cols[3].markdown(f'<span class="sim {"sim-hi" if sim>=80 else "sim-md" if sim>=65 else "sim-lo"}">{sim}%</span>', unsafe_allow_html=True)
                        row_cols[4].markdown(f'<span style="font-family:var(--mono)">{saran["kuantum"]:,}</span>', unsafe_allow_html=True)

                        fid_s = extract_fid(saran['surat_jalan'])
                        if fid_s:
                            row_cols[5].markdown(f'[🔗 Drive](https://drive.google.com/file/d/{fid_s}/view)')
                        else:
                            row_cols[5].markdown(f'[🔗 Buka]({saran["surat_jalan"]})')

                        # Preview
                        with row_cols[6]:
                            pa = st.session_state.saran_preview.get(f'item_{item_idx}') == s_idx_full
                            if st.button('👁️✕' if pa else '👁️', key=f'sprev_t3_{sk}', use_container_width=True):
                                st.session_state.saran_preview[f'item_{item_idx}'] = (None if pa else s_idx_full); st.rerun()

                        # Individual download
                        with row_cols[7]:
                            ls = saran['surat_jalan']
                            cs = st.session_state.dl_cache.get(ls)
                            if cs:
                                ext_s = infer_extension(cs)
                                fn_s = make_safe_filename(saran['nopol_f2'], saran['kuantum'], s_idx_full, ext_s)
                                st.download_button(f'⬇️ .{ext_s.upper()}', cs, fn_s, 'application/pdf' if ext_s=='pdf' else f'image/{ext_s}', key=f'sdl_c_{sk}', use_container_width=True)
                            else:
                                if st.button('⬇️ Unduh', key=f'sdl_t3_{sk}', use_container_width=True):
                                    with st.spinner('Mengunduh…'): ct_s = download_file(ls)
                                    if ct_s:
                                        st.session_state.dl_cache[ls] = ct_s; ext_s = infer_extension(ct_s)
                                        fn_s = make_safe_filename(saran['nopol_f2'], saran['kuantum'], s_idx_full, ext_s)
                                        st.download_button(f'💾 .{ext_s.upper()}', ct_s, fn_s, 'application/pdf' if ext_s=='pdf' else f'image/{ext_s}', key=f'sdl_s_{sk}', use_container_width=True)
                                    else: st.error('❌ Gagal.')

                        # Preview iframe
                        if st.session_state.saran_preview.get(f'item_{item_idx}') == s_idx_full:
                            purl_s = to_preview(saran['surat_jalan'])
                            if purl_s:
                                import streamlit.components.v1 as components
                                sc_c = '#157a3c' if sim>=80 else '#b45309' if sim>=65 else '#c41c1c'
                                st.markdown(f'<div class="alert a-sky" style="margin:8px 0"><span class="alert-ico">👁️</span><div>Preview — <b>{saran["nopol_f2"]}</b> · <b style="color:{sc_c}">{sim}%</b> · <a href="{purl_s}" target="_blank" style="font-size:.76rem">↗ Tab baru</a></div></div>', unsafe_allow_html=True)
                                components.html(f'<iframe src="{purl_s}" width="100%" height="680" style="border:1.5px solid #bae6fd;border-radius:11px;background:#fff" allow="autoplay"></iframe>', height=700)
                            else: st.error('Link tidak valid.')
                else:
                    st.caption(f'  {"🚫 Tidak ada saran untuk NOPOL ini." if item["kategori"]=="tidak_ada" else f"⬇️ Kurangi threshold ({min_sim}%) untuk melihat saran dengan kemiripan lebih rendah."}')

            # ── EXPORT CSV ───────────────────────────────────────────────
            st.markdown('<hr class="row-div">', unsafe_allow_html=True)
            if nopol_miss is not None and not nopol_miss.empty:
                cb_ex, _ = st.columns([2, 10])
                with cb_ex: st.download_button('📥 Export CSV', nopol_miss.to_csv(index=False).encode('utf-8'), 'nopol_tidak_ada.csv', 'text/csv', key='dl_b')
                st.dataframe(nopol_miss, use_container_width=True, hide_index=True)

    # ── TAB 4: SEMUA TIDAK MATCH ──────────────────────────────────────────
    with tab4:
        if missing.empty:
            st.markdown('<div class="empty-state"><div class="ico">🏆</div><h3>Sempurna! Semua Data Match!</h3><p>Setiap baris di File 1 berhasil dicocokkan dengan File 2.</p></div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="alert a-error"><span class="alert-ico">🔴</span><div><b>{n_all_miss} kombinasi tidak match</b> total.</div></div>', unsafe_allow_html=True)
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Total Tidak Match", n_all_miss); m2.metric("⚠️ Kuantum Beda", n_diff_k)
            m3.metric("❌ NOPOL Tidak Ada", n_miss_nopol); m4.metric("🔍 Ada Saran Mirip", n_nopol_mirip)
            all_m = missing.rename(columns={'nopol':'NOPOL','kuantum':'Kuantum File 1'}).copy()
            all_m['Kuantum File 1'] = all_m['Kuantum File 1'].astype(int)
            def get_ket(row):
                nopol = row['NOPOL']; f2m = df2_all[df2_all['nopol'] == nopol]
                if len(f2m) > 0:
                    ks = sorted(f2m['kuantum'].dropna().astype(int).unique().tolist())
                    d = ', '.join(map(str, ks[:5])) + (f' (+{len(ks)-5})' if len(ks)>5 else '')
                    return f'⚠️ Kuantum beda (di File 2: {d})'
                saran = find_nopol_suggestions(nopol, int(row['Kuantum File 1']), df2_all, top_n=3)
                if saran: top = saran[0]; return f'🔍 Saran: {top["nopol_f2"]} ({top["similarity"]}%)'
                return '❌ NOPOL tidak ada di File 2'
            with st.spinner('Menganalisis penyebab…'): all_m['Keterangan'] = all_m.apply(get_ket, axis=1)
            sa4c, _ = st.columns([3, 9])
            with sa4c: sa = st.text_input('', placeholder='🔍 Filter NOPOL…', label_visibility='collapsed', key='sa')
            if sa.strip(): all_m = all_m[all_m['NOPOL'].str.contains(re.escape(norm_nopol(sa)), na=False, case=False)].reset_index(drop=True)
            st.dataframe(all_m, use_container_width=True, hide_index=True)
            cde, _ = st.columns([2, 10])
            with cde: st.download_button('📥 Export CSV', all_m.to_csv(index=False).encode('utf-8'), 'semua_tidak_match.csv', 'text/csv', key='dl_c')

elif not st.session_state.processed:
    st.markdown('<div class="empty-state" style="margin-top:28px"><div class="ico">📂</div><h3>Upload File untuk Memulai</h3><p>Upload <b>File 1</b> (daftar target) dan <b>File 2</b> (database surat jalan) di atas, lalu klik <b>⚙️ Proses &amp; Cocokkan</b>.</p></div>', unsafe_allow_html=True)
